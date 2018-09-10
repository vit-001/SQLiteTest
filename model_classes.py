from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from gb_base import Base
from local_setting import fft


class Exercises:
    def __init__(self):
        self.base = dict()
        self.ex_base = {"Становая тяга с плинтов": {"pm": 220.0, "k": 1.0, "base": True},
                        "Жим штанги лежа на горизонтальной скамье": {"pm": 137.5, "k": 1.0, "base": True},
                        "Жим гантелей на скамье с наклоном вверх": {"pm": 39.6, "k": 2.0, "base": True},
                        "Жим штанги с груди сидя в смите": {"pm": 75.48, "k": 1.0, "base": True},
                        "Полуприсед со штангой": {"pm": 226.29, "k": 1.0, "base": True},
                        "Приседания со штангой на плечах": {"pm": 210.0, "k": 1.0, "base": True},
                        "Становая тяга": {"pm": 212.5, "k": 1.0, "base": True},
                        "Жим лежа в слинге": {"pm": 158.82, "k": 1.0, "base": True},
                        "Тяга штанги в наклоне": {"pm": 133.33, "k": 1.0, "base": False},
                        "Тяга гантели одной рукой в наклоне": {"pm": 61.88, "k": 2.0, "base": False},
                        "Французский жим лежа": {"pm": 60.0, "k": 1.0, "base": False},
                        "Становая тяга из ямы": {"pm": 200.0, "k": 1.0, "base": True},
                        "Вертикальная тяга верхнего блока к груди широким хватом": {"pm": 121.33, "k": 1.0,
                                                                                    "base": False},
                        "Подъем гантелей на бицепс сидя": {"pm": 19.58, "k": 2.0, "base": False},
                        "Горизонтальная тяга узким хватом в блочном тренажере": {"pm": 109.33, "k": 1.0, "base": False},
                        "Разгибание рук с гантелью из-за головы сидя": {"pm": 55.58, "k": 1.0, "base": False},
                        "Французский жим в блочном тренажере": {"pm": 90.0, "k": 1.0, "base": False},
                        "Подъем штанги на бицепс стоя": {"pm": 66.67, "k": 1.0, "base": False},
                        "Гиперэкстензия": {"pm": 57.6, "k": 1.0, "base": False},
                        "Бабочка": {"pm": 139.09, "k": 1.0, "base": False},
                        "Разведение гантелей в стороны стоя": {"pm": 23.69, "k": 2.0, "base": False},
                        "Жим книзу в блочном тренажере двумя руками": {"pm": 109.33, "k": 1.0, "base": False},
                        "Сгибание ног лежа на тренажере": {"pm": 104.73, "k": 1.0, "base": False},
                        "Жим П-грифа": {"pm": 98.18, "k": 1.0, "base": True},
                        "Жим штанги узким хватом лежа": {"pm": 95.62, "k": 1.0, "base": True},
                        "Пресс в тренажере": {"pm": 146.88, "k": 1.0, "base": False},
                        "Жим лежа с паузой": {"pm": 127.06, "k": 1.0, "base": True},
                        "Отжимания от скамьи в упоре сзади": {"pm": 0.0, "k": 1.0, "base": False},
                        "Приседания на скамью": {"pm": 164.57, "k": 1.0, "base": True},
                        "Жим гантелей сидя": {"pm": 36.93, "k": 2.0, "base": True},
                        "Жим штанги на скамье с наклоном вверх": {"pm": 120.0, "k": 1.0, "base": True},
                        "Жим гантелей лежа на горизонтальной скамье": {"pm": 44.83, "k": 2.0, "base": True},
                        "Жим ногами в тренажере": {"pm": 394.84, "k": 0.7, "base": True},
                        "Верхний хаммер": {"pm": 98.18, "k": 1.0, "base": False},
                        "Разгибания ног на тренажере": {"pm": 153.33, "k": 1.0, "base": False},
                        "Наклоны со штангой на плечах": {"pm": 173.79, "k": 1.0, "base": False},
                        "Махи сидя": {"pm": 16.88, "k": 1.0, "base": False},
                        "Нижний хаммер": {"pm": 66.67, "k": 1.0, "base": False},
                        "Подъем на бицепс в блочном тренажере стоя": {"pm": 86.67, "k": 1.0, "base": False},
                        "Присед с паузой внизу": {"pm": 152.73, "k": 1.0, "base": True},
                        "Сгибания ног сидя на тренажере": {"pm": 113.33, "k": 1.0, "base": False},
                        "Разведение гантелей сидя": {"pm": 14.73, "k": 2.0, "base": False},
                        "Жим с бруска": {"pm": 154.29, "k": 1.0, "base": True},
                        "Подъем гантелей перед собой": {"pm": 18.0, "k": 2.0, "base": False},
                        "Махи одной рукой поочередно": {"pm": 18.0, "k": 2.0, "base": False},
                        "Молоток": {"pm": 39.17, "k": 2.0, "base": False},
                        "Вертикальная тяга верхнего блока за голову широким хватом": {"pm": 78.67, "k": 1.0,
                                                                                      "base": False},
                        "Становая тяга с паузой ": {"pm": 174.86, "k": 1.0, "base": True},
                        "Жим сидя в тренажере": {"pm": 161.38, "k": 1.0, "base": False},
                        "Разведение гантелей на горизонтальной скамье": {"pm": 26.02, "k": 2.0, "base": False},
                        "Подьем гантелей перед собой сидя": {"pm": 12.96, "k": 2.0, "base": False},
                        "Скручивание на скамье с наклоном вниз": {"pm": 21.18, "k": 1.0, "base": False},
                        "Рычажная тяга": {"pm": 108.0, "k": 1.0, "base": False},
                        "Вертикальная тяга верхнего блока к груди D-рукоятки": {"pm": 89.33, "k": 1.0, "base": False},
                        "Подьем на бицепс на скамье скота": {"pm": 39.6, "k": 1.0, "base": False},
                        "Бицепс в тренажере": {"pm": 86.4, "k": 1.0, "base": False},
                        "Присед с двумя паузами": {"pm": 169.41, "k": 1.0, "base": True},
                        "Горизонтальная тяга широким хватом в блочном тренажере": {"pm": 105.12, "k": 1.0,
                                                                                   "base": False},
                        "Сгибание рук на бицепс в кроссовере": {"pm": 51.55, "k": 1.0, "base": False},
                        "Вертикальная тяга верхнего блока к груди обратным хватом": {"pm": 96.0, "k": 1.0,
                                                                                     "base": False},
                        "Отжимания на брусьях": {"pm": 27.24, "k": 1.0, "base": False},
                        "Протяжка на блоке": {"pm": 90.67, "k": 1.0, "base": False},
                        "Сведение в кроссовере через верхние блоки": {"pm": 76.24, "k": 1.0, "base": False},
                        "Жим Арнольда": {"pm": 19.58, "k": 2.0, "base": False},
                        "Отведение в кроссовере на среднюю дельту": {"pm": 26.67, "k": 1.0, "base": False},
                        "Жим в смите под углом": {"pm": 58.06, "k": 1.0, "base": False},
                        "Присед со штангой на груди": {"pm": 112.5, "k": 1.0, "base": True},
                        "Жим штанги стоя": {"pm": 58.06, "k": 1.0, "base": True},
                        "Жим штанги с груди сидя на опорной скамье": {"pm": 65.45, "k": 1.0, "base": True},
                        "Французский жим гантелей лежа": {"pm": 19.58, "k": 2.0, "base": False},
                        "Жим книзу одной рукой в блочном тренажере обратным хватом": {"pm": 36.0, "k": 2.0,
                                                                                      "base": False},
                        "Разведение ног на тренажере": {"pm": 192.71, "k": 1.0, "base": False},
                        "Жим с цепями": {"pm": 112.5, "k": 1.0, "base": True}}

        self.ex_base_old = [dict(name="Становая тяга с плинтов", pm=220.00, k=1.0, base=True),
                            dict(name="Жим штанги лежа на горизонтальной скамье", pm=137.50, k=1.0, base=True),
                            dict(name="Жим гантелей на скамье с наклоном вверх", pm=39.60, k=2.0, base=True),
                            dict(name="Жим штанги с груди сидя в смите", pm=75.48, k=1.0, base=True),
                            dict(name="Полуприсед со штангой", pm=226.29, k=1.0, base=True),
                            dict(name="Приседания со штангой на плечах", pm=210.00, k=1.0, base=True),
                            dict(name="Становая тяга", pm=212.50, k=1.0, base=True),
                            dict(name="Жим лежа в слинге", pm=158.82, k=1.0, base=True),
                            dict(name="Тяга штанги в наклоне", pm=133.33, k=1.0, base=False),
                            dict(name="Тяга гантели одной рукой в наклоне", pm=61.88, k=2.0, base=False),
                            dict(name="Французский жим лежа", pm=60.00, k=1.0, base=False),
                            dict(name="Становая тяга из ямы", pm=200.00, k=1.0, base=True),
                            dict(name="Вертикальная тяга верхнего блока к груди широким хватом", pm=121.33, k=1.0,
                                 base=False),
                            dict(name="Подъем гантелей на бицепс сидя", pm=19.58, k=2.0, base=False),
                            dict(name="Горизонтальная тяга узким хватом в блочном тренажере", pm=109.33, k=1.0,
                                 base=False),
                            dict(name="Разгибание рук с гантелью из-за головы сидя", pm=55.58, k=1.0, base=False),
                            dict(name="Французский жим в блочном тренажере", pm=90.00, k=1.0, base=False),
                            dict(name="Подъем штанги на бицепс стоя", pm=66.67, k=1.0, base=False),
                            dict(name="Гиперэкстензия", pm=57.60, k=1.0, base=False),
                            dict(name="Бабочка", pm=139.09, k=1.0, base=False),
                            dict(name="Разведение гантелей в стороны стоя", pm=23.69, k=2.0, base=False),
                            dict(name="Жим книзу в блочном тренажере двумя руками", pm=109.33, k=1.0, base=False),
                            dict(name="Сгибание ног лежа на тренажере", pm=104.73, k=1.0, base=False),
                            dict(name="Жим П-грифа", pm=98.18, k=1.0, base=True),
                            dict(name="Жим штанги узким хватом лежа", pm=95.62, k=1.0, base=True),
                            dict(name="Пресс в тренажере", pm=146.88, k=1.0, base=False),
                            dict(name="Жим лежа с паузой", pm=127.06, k=1.0, base=True),
                            dict(name="Отжимания от скамьи в упоре сзади", pm=0.00, k=1.0, base=False),
                            dict(name="Приседания на скамью", pm=164.57, k=1.0, base=True),
                            dict(name="Жим гантелей сидя", pm=36.93, k=2.0, base=True),
                            dict(name="Жим штанги на скамье с наклоном вверх", pm=120.00, k=1.0, base=True),
                            dict(name="Жим гантелей лежа на горизонтальной скамье", pm=44.83, k=2.0, base=True),
                            dict(name="Жим ногами в тренажере", pm=394.84, k=0.7, base=True),
                            dict(name="Верхний хаммер", pm=98.18, k=1.0, base=False),
                            dict(name="Разгибания ног на тренажере", pm=153.33, k=1.0, base=False),
                            dict(name="Наклоны со штангой на плечах", pm=173.79, k=1.0, base=False),
                            dict(name="Махи сидя", pm=16.88, k=1.0, base=False),
                            dict(name="Нижний хаммер", pm=66.67, k=1.0, base=False),
                            dict(name="Подъем на бицепс в блочном тренажере стоя", pm=86.67, k=1.0, base=False),
                            dict(name="Присед с паузой внизу", pm=152.73, k=1.0, base=True),
                            dict(name="Сгибания ног сидя на тренажере", pm=113.33, k=1.0, base=False),
                            dict(name="Разведение гантелей сидя", pm=14.73, k=2.0, base=False),
                            dict(name="Жим с бруска", pm=154.29, k=1.0, base=True),
                            dict(name="Подъем гантелей перед собой", pm=18.00, k=2.0, base=False),
                            dict(name="Махи одной рукой поочередно", pm=18.00, k=2.0, base=False),
                            dict(name="Молоток", pm=39.17, k=2.0, base=False),
                            dict(name="Вертикальная тяга верхнего блока за голову широким хватом", pm=78.67, k=1.0,
                                 base=False),
                            dict(name="Становая тяга с паузой ", pm=174.86, k=1.0, base=True),
                            dict(name="Жим сидя в тренажере", pm=161.38, k=1.0, base=False),
                            dict(name="Разведение гантелей на горизонтальной скамье", pm=26.02, k=2.0, base=False),
                            dict(name="Подьем гантелей перед собой сидя", pm=12.96, k=2.0, base=False),
                            dict(name="Скручивание на скамье с наклоном вниз", pm=21.18, k=1.0, base=False),
                            dict(name="Рычажная тяга", pm=108.00, k=1.0, base=False),
                            dict(name="Вертикальная тяга верхнего блока к груди D-рукоятки", pm=89.33, k=1.0,
                                 base=False),
                            dict(name="Подьем на бицепс на скамье скота", pm=39.60, k=1.0, base=False),
                            dict(name="Бицепс в тренажере", pm=86.40, k=1.0, base=False),
                            dict(name="Присед с двумя паузами", pm=169.41, k=1.0, base=True),
                            dict(name="Горизонтальная тяга широким хватом в блочном тренажере", pm=105.12, k=1.0,
                                 base=False),
                            dict(name="Сгибание рук на бицепс в кроссовере", pm=51.55, k=1.0, base=False),
                            dict(name="Вертикальная тяга верхнего блока к груди обратным хватом", pm=96.00, k=1.0,
                                 base=False),
                            dict(name="Отжимания на брусьях", pm=27.24, k=1.0, base=False),
                            dict(name="Протяжка на блоке", pm=90.67, k=1.0, base=False),
                            dict(name="Сведение в кроссовере через верхние блоки", pm=76.24, k=1.0, base=False),
                            dict(name="Жим Арнольда", pm=19.58, k=2.0, base=False),
                            dict(name="Отведение в кроссовере на среднюю дельту", pm=26.67, k=1.0, base=False),
                            dict(name="Жим в смите под углом", pm=58.06, k=1.0, base=False),
                            dict(name="Присед со штангой на груди", pm=112.50, k=1.0, base=True),
                            dict(name="Жим штанги стоя", pm=58.06, k=1.0, base=True),
                            dict(name="Жим штанги с груди сидя на опорной скамье", pm=65.45, k=1.0, base=True),
                            dict(name="Французский жим гантелей лежа", pm=19.58, k=2.0, base=False),
                            dict(name="Жим книзу одной рукой в блочном тренажере обратным хватом", pm=36.00, k=2.0,
                                 base=False),
                            dict(name="Разведение ног на тренажере", pm=192.71, k=1.0, base=False),
                            dict(name="Жим с цепями", pm=112.50, k=1.0, base=True)]

    def add(self, name: str, value: float, repeat: int):
        old = self.base.get(name, {'max': 0.0, 'repeat': 1})
        if value < old['max']:
            return

        if value > old['max']:
            old['max'] = value
            old['repeat'] = repeat
        else:
            if repeat > old['repeat']:
                old['repeat'] = repeat
            else:
                return
        self.base[name] = old

    def output(self):
        for name, data in self.base.items():
            m = data['max']
            k = data['repeat']
            # print(name, m, 'x', k, '%1.1f' % (m * (36 / (37 - k))))
            print(f'dict(name="{name}", pm={(m * (36 / (37 - k))):.2f}, k=1.0, base=True),')

        print(self.ex_base_old)

    def get_1PM(self, name: str) -> float:
        ex = self.base.get(name, {'max': 0.0, 'repeat': 1})
        m = ex['max']
        k = ex['repeat']
        return (m * (36 / (37 - k)))

    def _get_1PM(self, name: str) -> float:
        pass

    def is_base(self, ex_name:str)->bool:
        try:
            return self.ex_base[ex_name]['base']
        except KeyError:
            return False


    def _convert_base(self):
        for item in self.ex_base_old:
            # print(item)

            print(f'"{item["name"]}" : {{"pm": {item["pm"]}, "k": {item["k"]}, "base": {item["base"]} }},')


class Statistic:
    max_repeat = 17
    exercises = Exercises()

    def stat(self):
        print('Not implemented')

    @staticmethod
    def proceed(base: Base):
        for workout in base.days:
            for ex in workout.excercises:
                # ex.stat()
                for item in ex.sets:
                    Statistic.exercises.add(ex.name, item.weight, item.repeats)

        Statistic.exercises.output()

    @staticmethod
    def get_1pm(ex_name):
        return Statistic.exercises.get_1PM(ex_name)

    @staticmethod
    def is_base(ex_name:str)->bool:
        return Statistic.exercises.is_base(ex_name)


class CreateTable:
    @staticmethod
    def proceed(base: Base, fd):
        for workout in base.days:
            print(workout.date, ';', workout.name, file=fd)
            for ex in workout.excercises:
                table_width = Statistic.max_repeat + 3
                first_line = ['' for i in range(table_width)]
                second_line = ['' for i in range(table_width)]
                third_line = ['' for i in range(table_width)]
                use_third_line = False

                first_line[2] = ex.name
                i = 3

                for set in ex.sets:
                    first_line[i] = fft(set.weight)
                    second_line[i] = '%10d' % set.repeats

                    if set.anydata:
                        third_line[i] = set.anydata
                        use_third_line = True
                    if set.comment:
                        third_line[i] = third_line[i] + set.comment
                        use_third_line = True

                    i = i + 1

                for item in first_line:
                    print(item, end=';', file=fd)

                if ex.sets:
                    print('V;КПШ;Ио;ср вес; 1ПМ; КО', file=fd)

                    for item in second_line:
                        print(item, end=';', file=fd)

                    print('"=СУММПРОИЗВ(R[-1]C[-17]:R[-1]C[-1];RC[-17]:RC[-1])"', end=';', file=fd)
                    print('"=СУММ(RC[-18]:RC[-2])"', end=';', file=fd)
                    print('"=RC[1]/RC[2]"', end=';', file=fd)
                    print('"=RC[-3]/RC[-2]"', end=';', file=fd)
                    print(fft(Statistic.exercises.get_1PM(ex.name)), end=';', file=fd)
                    print('"=RC[-3]*RC[-4]"', file=fd)

                    if use_third_line:
                        for item in third_line:
                            print(item, end=';', file=fd)
                        print(file=fd)

                else:
                    print(file=fd)


class ExportXLS:
    def __init__(self, base: Base):
        self.base = base

    def proceed_to_xls(self, fname):
        wb = Workbook()
        ws = wb.active

        c_row = 1

        for workout in self.base.days:
            day_row = c_row
            c_row = c_row + 1

            ws.cell(day_row, 1).value = workout.date
            ws.cell(day_row, 2).value = workout.name
            ws.cell(day_row, 1).font = Font(bold=True)
            ws.cell(day_row, 2).font = Font(bold=True)

            for ex in workout.excercises:
                c_row = c_row + 1
                if Statistic.is_base(ex.name):
                    c_row = self.full_exercise(ws,ex,c_row)
                else:
                    c_row = self.short_exercise(ws,ex,c_row)

            self.bold_center(ws.cell(day_row, 21), 'V')
            self.bold_center(ws.cell(day_row, 22), 'КПШ')
            self.bold_center(ws.cell(day_row, 23), 'Ио')
            self.bold_center(ws.cell(day_row, 24), 'ср вес')
            self.bold_center(ws.cell(day_row, 25), '1ПМ')
            self.bold_center(ws.cell(day_row, 26), 'КО')
            self.bold_center(ws.cell(day_row + 1, 21), f'=SUM(U{day_row+2}:U{c_row})')
            self.bold_center(ws.cell(day_row + 1, 22), f'=SUM(V{day_row+2}:V{c_row})')
            self.bold_center(ws.cell(day_row + 1, 26), f'=SUM(Z{day_row+2}:Z{c_row})')
            ws.cell(day_row + 1, 26).number_format = '0.0'

            c_row = c_row + 2

        self.format_exersises_sheet(ws, c_row)
        wb.save(fname)

    def full_exercise(self, ws, ex, row:int)-> int:
        use_third_line = False

        ws.cell(row, 3).value = ex.name
        ws.cell(row, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        i = 4
        for set in ex.sets:
            ws.cell(row, i).value = set.weight
            ws.cell(row + 2, i).value = set.repeats
            if Statistic.exercises.get_1PM(ex.name) > 0.01:
                ws.cell(row + 1, i).value = set.weight / Statistic.get_1pm(ex.name)
                ws.cell(row + 1, i).number_format = '0.0%'

            t = ''
            if set.anydata:
                t = t + set.anydata
                use_third_line = True
            if set.comment:
                t = t + set.comment
                use_third_line = True

            ws.cell(row + 3, i).value = t

            i = i + 1

        if ex.sets:

            # ws.merge_cells(f'C{c_row}:C{c_row+2}')

            ws.cell(row, 21).value = f'=SUMPRODUCT((D{row}:T{row}>Y{row}/2.01)*1,D{row}:T{row},D{row+2}:T{row+2})'        #=СУММПРОИЗВ((D400:T400>Y400/2)*1;D402:T402;D400:T400)
            ws.cell(row, 22).value = f'=SUM(D{row+2}:T{row+2})'
            ws.cell(row, 23).value = f'=X{row}/Y{row}'
            ws.cell(row, 24).value = f'=U{row}/V{row}'
            ws.cell(row, 25).value = Statistic.get_1pm(ex.name)
            ws.cell(row, 26).value = f'=W{row}*V{row}'
            ws.cell(row, 23).number_format = '0.0%'
            ws.cell(row, 24).number_format = '0.0'
            ws.cell(row, 25).number_format = '0.0'
            ws.cell(row, 26).number_format = '0.0'

            row = row + 2

            if use_third_line:
                row = row + 1
        return row

    def short_exercise(self, ws, ex, row:int)-> int:
        ws.cell(row, 3).value = ex.name
        ws.cell(row, 3).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        i = 4
        for set in ex.sets:
            ws.cell(row, i).value = set.repeats
            i = i + 1
        return row

    def bold_center(self, cell, value):
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='top')

    def format_exersises_sheet(self, ws, last_row):

        align_left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

        # for col in ws.iter_cols(min_col=21, max_col=26, max_row=last_row):
        #     for cell in col:
        #         cell.alignment=align_center

        # for cell in ws['C']:
        #     cell.alignment = align_left_wrap

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 1
        ws.column_dimensions['C'].width = 30


if __name__ == "__main__":
    ex = Exercises()
    ex._convert_base()
